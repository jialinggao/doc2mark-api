import openpyxl
import xlrd
import io
import os
import tempfile
import subprocess
import time
from typing import BinaryIO, List, Optional, Tuple
from loguru import logger
from app.models import ImageMode
from app.services.general_converter import general_converter


def clean_word_special_chars(markdown: str) -> str:
    """
    清理 Word 文档中的特殊字符

    Word 文档常包含 Unicode 专用区 (Private Use Area, 0xE000-0xF900)
    的字符，这些字符在 Markdown 中通常没有实际意义，需要清理。

    Args:
        markdown: 原始 Markdown 文本

    Returns:
        清理后的 Markdown 文本
    """
    cleaned = markdown

    # 清理 Unicode 专用区字符 (0xE000-0xF900)
    for code in range(0xE000, 0xF900):
        char = chr(code)
        if char in cleaned:
            cleaned = cleaned.replace(char, '')

    return cleaned


class ExcelConverter:
    """
    Excel 专用转换器 - 正确处理特殊符号、合并单元格和 NaN 值

    主要功能：
    1. 支持 .xls (Excel 97-2003) 和 .xlsx (Excel 2007+) 格式
    2. 正确保留特殊符号（如 √、① 等）
    3. 正确处理合并单元格
    4. 清理空值 (NaN/None)
    5. 生成整洁的 Markdown 表格
    """

    def convert(
        self,
        file_stream: BinaryIO,
        filename: str,
        enable_ocr: bool = False,
        enable_llm: bool = False,
        image_mode: ImageMode = ImageMode.BASE64,
        image_quality: int = 100,
        max_image_size: int = -1
    ) -> dict:
        """
        转换 Excel 文件为 Markdown

        先用专用 Excel 解析器处理，如果失败则回退到通用处理

        Args:
            file_stream: 文件二进制流
            filename: 文件名（用于判断文件格式）
            enable_ocr: 是否启用 OCR（Excel 转换中暂不使用）
            enable_llm: 是否启用 LLM（Excel 转换中暂不使用）
            image_mode: 图片输出模式（Excel 转换中暂不使用）
            image_quality: 图片质量（Excel 转换中暂不使用）
            max_image_size: 图片最大尺寸（Excel 转换中暂不使用）

        Returns:
            转换结果字典，包含:
            - filename: 文件名
            - markdown: Markdown 文本
            - images: 图片列表（Excel 转换中为空）
            - duration: 转换耗时（秒）
        """
        try:
            return self._convert_excel_internal(file_stream, filename, enable_ocr, enable_llm, image_mode, image_quality, max_image_size)
        except Exception as e:
            logger.warning(f"[Excel 处理] 专用处理失败，回退到通用处理: {e}")
            file_stream.seek(0)
            return general_converter.convert(file_stream, filename, enable_ocr, enable_llm, image_mode, image_quality, max_image_size)

    def _convert_excel_internal(
        self,
        file_stream: BinaryIO,
        filename: str,
        enable_ocr: bool = False,
        enable_llm: bool = False,
        image_mode: ImageMode = ImageMode.BASE64,
        image_quality: int = 100,
        max_image_size: int = -1
    ) -> dict:
        """
        内部 Excel 转换实现
        """
        start_time = time.time()

        # 获取文件扩展名
        ext = os.path.splitext(filename.lower())[1]

        # 根据扩展名选择解析方法
        if ext == '.xls':
            markdown_text = self._parse_xls(file_stream, filename)
        elif ext == '.xlsx':
            markdown_text = self._parse_xlsx(file_stream)
        else:
            markdown_text = self._parse_xlsx(file_stream)

        duration = time.time() - start_time

        return {
            "filename": filename,
            "markdown": clean_word_special_chars(markdown_text),
            "images": [],
            "duration": round(duration, 2)
        }

    def _parse_xls(self, file_stream: BinaryIO, filename: str) -> str:
        """
        解析 .xls (Excel 97-2003) 格式文件

        优先尝试用 LibreOffice 转换为 .xlsx，失败时回退到 xlrd 直接读取

        Args:
            file_stream: 文件二进制流
            filename: 文件名

        Returns:
            Markdown 文本
        """
        # 创建临时目录处理文件
        with tempfile.TemporaryDirectory() as tmpdir:
            # 保存文件到临时目录
            input_path = os.path.join(tmpdir, filename)
            with open(input_path, "wb") as f:
                f.write(file_stream.read())

            # 构建 LibreOffice 转换命令
            cmd = [
                "libreoffice", "--headless", "--convert-to", "xlsx",
                "--outdir", tmpdir, input_path
            ]

            logger.info(f"[Excel处理] 将.xls转换为.xlsx")
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)

            # 检查转换是否成功
            if result.returncode != 0:
                logger.warning(f"[Excel处理] LibreOffice转换失败，尝试用xlrd直接读取: {result.stderr}")
                return self._parse_xls_with_xlrd(input_path)

            # 检查转换后的文件是否存在
            base_name = os.path.splitext(filename)[0]
            xlsx_path = os.path.join(tmpdir, base_name + ".xlsx")

            if not os.path.exists(xlsx_path):
                logger.warning(f"[Excel处理] 转换后的xlsx文件不存在，尝试用xlrd直接读取")
                return self._parse_xls_with_xlrd(input_path)

            # 使用 openpyxl 解析转换后的 .xlsx
            with open(xlsx_path, "rb") as f:
                return self._parse_xlsx(f)

    def _parse_xls_with_xlrd(self, file_path: str) -> str:
        """
        使用 xlrd 直接解析 .xls 文件（当 LibreOffice 转换失败时使用）

        Args:
            file_path: 文件路径

        Returns:
            Markdown 文本
        """
        # 打开工作簿，保留格式信息
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        all_md = []

        # 遍历所有工作表
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            all_md.append(f"## {sheet_name}\n")

            # 读取单元格数据
            data = []
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    val = self._format_xlrd_cell(cell)
                    row_data.append(val)
                data.append(row_data)

            # 填充合并单元格、裁剪空值、转换为 Markdown 表格
            data = self._fill_merged_cells_xlrd(sheet, data)
            data = self._trim_empty_rows_cols(data)
            md_table = self._data_to_markdown_table(data)
            all_md.append(md_table)

        return "\n\n".join(all_md)

    def _format_xlrd_cell(self, cell) -> str:
        """
        格式化 xlrd 读取的单元格值

        Args:
            cell: xlrd 单元格对象

        Returns:
            格式化后的字符串
        """
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return ""
        elif cell.ctype == xlrd.XL_CELL_TEXT:
            return str(cell.value).strip()
        elif cell.ctype == xlrd.XL_CELL_NUMBER:
            if cell.value == int(cell.value):
                return str(int(cell.value))
            return str(cell.value).strip()
        elif cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return str(xlrd.xldate.xldate_as_datetime(cell.value, 0))
            except:
                return str(cell.value).strip()
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return "√" if cell.value else ""
        else:
            val = str(cell.value).strip()
            if val in ['nan', 'NaN', 'None', '']:
                return ""
            return val

    def _fill_merged_cells_xlrd(self, sheet, data: List[List[str]]) -> List[List[str]]:
        """
        填充 xlrd 解析的合并单元格

        将合并单元格区域的值填充到所有合并的单元格中

        Args:
            sheet: xlrd 工作表对象
            data: 原始数据

        Returns:
            填充合并单元格后的数据
        """
        merged_cells = sheet.merged_cells
        for (rlo, rhi, clo, chi) in merged_cells:
            top_left_val = data[rlo][clo]
            for row_idx in range(rlo, rhi):
                for col_idx in range(clo, chi):
                    data[row_idx][col_idx] = top_left_val
        return data

    def _parse_xlsx(self, file_stream: BinaryIO) -> str:
        """
        解析 .xlsx (Excel 2007+) 格式文件

        Args:
            file_stream: 文件二进制流

        Returns:
            Markdown 文本
        """
        file_stream.seek(0)
        # 加载工作簿，仅读取数据（不读取公式）
        wb = openpyxl.load_workbook(
            io.BytesIO(file_stream.read()),
            data_only=True,
            read_only=False
        )

        all_md = []

        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_md.append(f"## {sheet_name}\n")

            max_row = ws.max_row
            max_col = ws.max_column

            # 读取单元格数据
            data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                row_data = []
                for cell in row:
                    val = self._format_openpyxl_cell(cell)
                    row_data.append(val)
                data.append(row_data)

            # 填充合并单元格、裁剪空值、转换为 Markdown 表格
            data = self._fill_merged_cells_openpyxl(ws, data)
            data = self._trim_empty_rows_cols(data)

            if data:
                md_table = self._data_to_markdown_table(data)
                all_md.append(md_table)

        return "\n\n".join(all_md)

    def _format_openpyxl_cell(self, cell) -> str:
        """
        格式化 openpyxl 读取的单元格值

        Args:
            cell: openpyxl 单元格对象

        Returns:
            格式化后的字符串
        """
        val = cell.value

        if val is None:
            return ""

        if isinstance(val, bool):
            return "√" if val else ""

        if isinstance(val, (int, float)):
            if isinstance(val, float) and val == int(val):
                return str(int(val))
            return str(val).strip()

        val_str = str(val).strip()

        if val_str in ['nan', 'NaN', 'None', '']:
            return ""

        return val_str.replace('\n', '<br>')

    def _fill_merged_cells_openpyxl(self, ws, data: List[List[str]]) -> List[List[str]]:
        """
        填充 openpyxl 解析的合并单元格

        Args:
            ws: openpyxl 工作表对象
            data: 原始数据

        Returns:
            填充合并单元格后的数据
        """
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds

            if min_row <= len(data) and min_col <= len(data[min_row - 1]):
                top_left_val = data[min_row - 1][min_col - 1]

                for row_idx in range(min_row - 1, min(max_row, len(data))):
                    for col_idx in range(min_col - 1, min(max_col, len(data[row_idx]))):
                        data[row_idx][col_idx] = top_left_val

        return data

    def _trim_empty_rows_cols(self, data: List[List[str]]) -> List[List[str]]:
        """
        裁剪表格的空行和空列

        移除顶部和底部的空行，以及右侧的空列

        Args:
            data: 原始数据

        Returns:
            裁剪后的数据
        """
        if not data:
            return data

        # 移除底部的空行
        while data and all(cell == "" for cell in data[-1]):
            data.pop()

        # 移除顶部的空行
        while data and all(cell == "" for cell in data[0]):
            data.pop(0)

        if not data:
            return data

        # 找到最右侧非空列
        max_col = len(data[0])
        while max_col > 0 and all(row[max_col - 1] == "" for row in data if len(row) >= max_col):
            max_col -= 1

        # 裁剪到实际内容区域
        trimmed = []
        for row in data:
            trimmed_row = row[:max_col]
            while len(trimmed_row) < max_col:
                trimmed_row.append("")
            trimmed.append(trimmed_row)

        return trimmed

    def _data_to_markdown_table(self, data: List[List[str]]) -> str:
        """
        将二维列表数据转换为 Markdown 表格格式

        Args:
            data: 二维列表数据

        Returns:
            Markdown 表格字符串
        """
        if not data or not data[0]:
            return ""

        num_cols = len(data[0])

        md_lines = []

        # 表头
        header = data[0]
        md_lines.append("| " + " | ".join(self._escape_table_cell(c) for c in header) + " |")
        # 分隔线
        md_lines.append("| " + " | ".join(["---"] * num_cols) + " |")

        # 数据行
        for row in data[1:]:
            while len(row) < num_cols:
                row.append("")
            md_lines.append("| " + " | ".join(self._escape_table_cell(c) for c in row) + " |")

        return "\n".join(md_lines)

    def _escape_table_cell(self, text: str) -> str:
        """
        转义 Markdown 表格单元格中的特殊字符

        主要处理:
        - 管道符 | → 转义为 \|
        - 换行符 → 替换为 <br>

        Args:
            text: 原始文本

        Returns:
            转义后的文本
        """
        if text is None:
            return ""
        text = str(text)
        text = text.replace("|", "\\|")
        text = text.replace("\n", "<br>")
        return text


# Excel 转换器单例实例
excel_converter = ExcelConverter()
